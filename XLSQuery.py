from tkinter import *
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from tkinter import ttk
import re
import sqlite3
from datetime import datetime
from tkinter import messagebox
# import pandas as pd

# df = pd.read_csv('archivo.csv')
# print(df)


import os
import xlrd
# def info_data_type(dato):
#     # Intentar convertir a entero
#     try:
#         int(dato)
#         return "INTEGER"
#     except ValueError:
#         pass

#     # Intentar convertir a decimal
#     try:
#         float(dato)
#         return "REAL"
#     except ValueError:
#         pass

#     # Lista de formatos de fecha posibles
#     formatos_fecha = [
#         "%d-%b-%y %H:%M:%S",  # 26-Feb-25 16:21:48
#         "%d-%b-%y",           # 26-Feb-25
#         "%Y-%m-%d %H:%M:%S",  # 2025-02-26 16:21:48
#         "%Y-%m-%d",           # 2025-02-26
#         "%d/%m/%Y %H:%M:%S",  # 26/02/2025 16:21:48
#         "%d/%m/%Y",           # 26/02/2025
#         "%m/%d/%Y %H:%M:%S",  # 02/26/2025 16:21:48
#         "%m/%d/%Y",           # 02/26/2025
#         "%d-%m-%Y %H:%M:%S",  # 26-02-2025 16:21:48
#         "%d-%m-%Y"            # 26-02-2025
#     ]
    
#     # Intentar convertir a fecha
#     for formato in formatos_fecha:
#         try:
#             fecha_obj = datetime.strptime(dato, formato)
#             if "%H" in formato:  # Si el formato incluye hora
#                 return "DATETIME"
#             else:
#                 return "DATE"
#         except ValueError:
#             pass
    
#     return "TEXT"
# def convertir_fecha(fecha_str):
#     # Lista de formatos posibles
#     formatos = [
#         "%d-%b-%y %H:%M:%S",  # 26-Feb-25 16:21:48
#         "%d-%b-%y",           # 26-Feb-25
#         "%Y-%m-%d %H:%M:%S",  # 2025-02-26 16:21:48
#         "%Y-%m-%d",           # 2025-02-26
#         "%d/%m/%Y %H:%M:%S",  # 26/02/2025 16:21:48
#         "%d/%m/%Y",           # 26/02/2025
#         "%m/%d/%Y %H:%M:%S",  # 02/26/2025 16:21:48
#         "%m/%d/%Y",           # 02/26/2025
#         "%d-%m-%Y %H:%M:%S",  # 26-02-2025 16:21:48
#         "%d-%m-%Y"            # 26-02-2025
#     ]    
#     for formato in formatos:
#         try:
#             fecha_obj = datetime.strptime(fecha_str, formato)  # Intenta convertir
#             if "%H" in formato:  # Si el formato original tenía hora
#                 return fecha_obj.strftime("%Y-%m-%d %H:%M:%S")
#             else:  # Si no tenía hora
#                 return fecha_obj.strftime("%Y-%m-%d")
#         except ValueError:
#             pass  # Si falla, intenta con otro formato
    
#     return "Formato de fecha no reconocido"  # Si ningún formato coincide

# Pruebas
# fechas = [
#     "26-Feb-25 16:21:48",
#     "26-Feb-25",
#     "2025-02-26 16:21:48",
#     "2025-02-26",
#     "26/02/2025 16:21:48",
#     "26/02/2025",
#     "02/26/2025 16:21:48",
#     "02/26/2025",
#     "26-02-2025 16:21:48",
#     "26-02-2025"
# ]

# for fecha in fechas:
#     print(f"{fecha} → {convertir_fecha(fecha)}")

# VARIABLES Y CONSTANTES
name_path_file = ''
row_muestra = None
row_title = None
workbook  = None
extension_user = None
URL_DATABASE = None
DATA_EXECUTE = None
# f_columns = None
# Lista de palabras reservadas de SQL
palabras_reservadas = [
    "with","select", "insert", "update", "delete", "from", "where", "join", "into", "drop", "alter", "create",
    "table", "column", "values", "as", "and", "or", "not", "is", "in", "like", "between", "group", "having",
    "order", "by", "distinct", "union", "left", "right", "inner", "outer", "exists", "case", "when", "then",
    "else", "end", "null", "true", "false", "on", "between", "like", "limit", "offset", "primary", "foreign",
    "key", "check", "constraint","count","IF","TEMP",";",",","=","DATE","DATE","DATETIME","hours","day","now"
]

def read_excel_file():
    pass
def name_validate_sql(name):
        # 1. Verificar que el nombre no esté vacío
        if not name:
            return False, f"[{name}] no debe ser un valor nulo"
        # 2. Verificar que el nombre no comience con un número
        if name[0].isdigit():
            return False, f"[{name}] no puede comenzar con un número."
        # 3. Verificar que el nombre no contenga espacios
        if " " in name:
            return False, f"[{name}] contiene espacios"
        # 4. Verificar que el nombre no contenga caracteres especiales no permitidos
        if not re.match("^[a-zA-Z0-9_]+$", name):
            return False, f"[{name}] solo debe contener letras, números(nunca al inicio del nombre) y guiones bajos, elimina los caracteres no permitidos"
        # 5. Verificar que el nombre no sea una palabra reservada
        if name.lower() in palabras_reservadas:
            return False, f"No se puede usar [{name}] porque es una palabra reservada"
        return True, f"[{name}] OK"
def save_data(file_path, t_name, sheet_name, all_tabs, sufijo): #name_path_file, table_name.get(), list_sheets.get(),var.get()
    # Abre el diálogo para seleccionar una carpeta
    # carpeta = filedialog.askdirectory(title="Selecciona una carpeta")
    if var_saved.get(): 
        file_path = "DATA/data_main.db"
    else:
        file_path = filedialog.asksaveasfilename(
            title="Guardar base de datos SQLite como...",
            defaultextension=".db",  # Extensión por defecto si el usuario no pone una
            filetypes=[
                ("SQLite Database (*.db, *.sqlite, *.sqlite3, *.db3, *.sdb, *.sl3)",
                "*.sqlite *.sqlite3 *.db *.db3 *.sdb *.sl3"),
                ("Todos los archivos", "*.*")
            ]
        )
    if not(file_path):
        messagebox.showwarning("Alerta", 'Debe seleccionar una ruta para almacenar la base de datos')
        return
    # if carpeta:
    def SQL_INSERT_DATA(list_sheet_names_ok,name_from_edittable): #name_from_edittable existe una lista con un unico elemento cuando el usuario elige solo una pestaña del libro excel
        sufijo_table = sufijo.replace(" ","")
        if len(sufijo_table):
            if not re.match("^[a-zA-Z0-9_]+$", sufijo_table):
                return messagebox.showinfo("Alerta", f"El sufijo [{sufijo_table}] solo debe contener letras, números y guiones bajos, elimina los demas caracteres")
            
        print("HOJAS VALIDADAS",list_sheet_names_ok)
        # Conectar a la base de datos (si no existe, se crea)
        # conexion = sqlite3.connect(f'{carpeta}/data_main.db')
        conexion = sqlite3.connect(file_path)
        cursor = conexion.cursor()
        for sheet_item in list_sheet_names_ok:
            
            sheet_browser = sheet_item
            
            #Cambia el nombre personalizado de la tabla si existe
            if len(name_from_edittable):
                sheet_item = name_from_edittable[0]
            
            sheet_item = f"{sheet_item.lower()}{sufijo_table}"

            print("Crear tabla:", sheet_item)
            # Consulta para verificar existencia
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (sheet_item,))
            existe = cursor.fetchone() is not None
            
            if existe:
                print(f"La tabla '{sheet_item}' existe.")
                # conexion.close()
                messagebox.showwarning("Alerta", f'La tabla "{sheet_item}" será omitida porque ya existe')
            else:
                try:
                    if extension_user == '.xlsx':
                        sheet_now = workbook[sheet_browser]
                        # Leer la primera fila (títulos)
                        #row_title_sheet = [cell.value for cell in sheet_now[1]]
                        row_title_sheet = [cell.value for cell in sheet_now[1] if cell.value and str(cell.value).strip() != ""]
                        nrows = sheet_now.max_row
                        if not(nrows > 1): 
                            raise ValueError(f"La Hoja '{sheet_browser}' esta vacia")
                        else:
                            lista_muestra = [cell.value for cell in sheet_now[2]]
                            row_muestra_sheet = lista_muestra[:len(row_title_sheet)]

                    elif extension_user == '.xls':
                        sheet_now = workbook.sheet_by_name(sheet_browser)
                        row_title_sheet = sheet_now.row_values(0)
                        nrows = sheet_now.nrows
                        if not(nrows > 1): 
                            # print(f"La Hoja {sheet_browser} esta vacia")
                            raise ValueError(f"La Hoja '{sheet_browser}' esta vacia")
                        else:
                            lista_muestra = sheet_now.row_values(1)
                            row_muestra_sheet = lista_muestra[:len(row_title_sheet)]
                    else:
                        raise ValueError("Formato de archivo no soportado: usa .xls o .xlsx")
                    
                    
                    # row_muestra_sheet = sheet_now.row_values(1)
                    # Trabaja con la fila aquí
                    estructura_sql = ''
                    estructura_sql_columns = ''
                    estructura_sql_columns_signo = ''
                    
                    #EXTRAE DATOS DE LAS ENTRADAS DE TEXTO
                    column_muestra = 0
                    print(sheet_item,row_muestra_sheet)
                    for col in row_muestra_sheet:
                        data_column = row_muestra_sheet[column_muestra]
                        data_type = None
                        if  data_column == None or type(data_column) == str or isinstance(data_column, datetime):
                            data_type = "TEXT"
                        elif  type(data_column) == int or type(data_column) == float:
                            data_type = "REAL"
                        
                        estructura_sql+= f"{row_title_sheet[column_muestra]} {data_type}"
                        estructura_sql_columns+= f"{row_title_sheet[column_muestra]}"
                        estructura_sql_columns_signo+= "?"
                        if column_muestra+1 < len(row_title_sheet):
                            estructura_sql+=", "
                            estructura_sql_columns+=", "
                            estructura_sql_columns_signo+=", "
                        print("Column: ", row_title_sheet[column_muestra],col, data_type)
                        column_muestra+=1
                    # Crear la tabla
                    cursor.execute(f'''CREATE TABLE IF NOT EXISTS {sheet_item} ({estructura_sql})''')
                    print(f"CREATE TABLE IF NOT EXISTS {sheet_item} ({estructura_sql})")
                    conexion.commit()
                    
                    #==================================================================
                    # INSERTAR LOS REGISTROS EN LA NUEVA TABLA
                    #==================================================================
                    first_sheet = sheet_now
                    
                    
                    data_insert = []


                    if extension_user == '.xlsx':
                        # Iterar desde la segunda fila (saltando encabezados)
                        for row in first_sheet.iter_rows(min_row=2, values_only=True):
                            data_insert.append(list(row[:len(row_title_sheet)]))
                    elif extension_user == '.xls':
                        for row_idx in range(1, first_sheet.nrows):  # Salta la fila de encabezados
                            fila = first_sheet.row_values(row_idx)
                            data_insert.append(fila)
                    else:
                        raise ValueError("Formato de archivo no soportado: usa .xls o .xlsx")
                    
                    print(f"INSERT INTO {sheet_item} ({estructura_sql_columns}) VALUES ({estructura_sql_columns_signo})")
                    cursor.executemany(f"INSERT INTO {sheet_item} ({estructura_sql_columns}) VALUES ({estructura_sql_columns_signo})", data_insert)
                    conexion.commit()
                except ValueError  as e:
                    print(e)
                    messagebox.showinfo("Alerta", e)

        conexion.close()
        # Muestra la notificación
        print(r"""
  _____        _        _                              
 |  __ \      | |      | |                             
 | |  | | __ _| |_ __ _| |__   __ _ ___  ___  
 | |  | |/ _` | __/ _` | '_ \ / _` / __|/ _ \_
 | |__| | (_| | || (_| | |_) | (_| \__ \  __/
 |_____/ \__,_|\__\__,_|_.__/ \__,_|___/\___|    

    ( )   ___ 
    ||  / __|
    ||  \__ \_
    ||  |___/
                                                 
  _____  ______   ____   ____   __         
 |  __ \|  ____| | __ \ |  _  \ \ \   / /         
 | |__) | |__   | |__\ || | \ |  \ \_/ /                 
 |  _  /|  __|  |  __  || | | |   \   /                  
 | | \ \| |____ | |  | || |_| |   |  |                   
 |_|  \_\______|| |  | |\____/   |__|                   

                  D A T A B A S E   R E A D Y
                                Powered  by Breyner J.
""")
        if messagebox.showinfo("Proceso Terminado", "Ya puedes realizar consultas SQL"):
            for widget in f_columns.winfo_children():
                widget.destroy()
        
        refrescar()
        lbl_file_path_out.delete(0, tk.END)      # Borrar todo el contenido
        lbl_file_path_out.insert(0, file_path)  # Insertar texto desde el inicio
    if all_tabs:
        #LISTA DE HOJAS VALIDADAS
        SHEET_OK_INSERT = []
        # if extension_user == '.xls':
        table_error_name = []
        column_name_error = []
        if extension_user == '.xlsx':
            hojas = workbook.sheetnames
        elif extension_user == '.xls':
            hojas = workbook.sheet_names()
        else:
            raise ValueError("Formato de archivo no soportado: usa .xls o .xlsx")

        for hoja in hojas:
            sheet_table_name = hoja.lower()
            validate_name = name_validate_sql(sheet_table_name)
            if validate_name[0]:
                print("Tabla lista para ser insertada",sheet_table_name)
                #SELECCIONA LA HOJA ACTUAL EN ITERACION Y OBTIENE LA CABECERA O PRIMERA FILA
                if extension_user == '.xlsx':
                    sheet = workbook[hoja]
                    # Leer la primera fila (títulos)
                    #row_title = [cell.value for cell in sheet[1]]
                    row_title = [cell.value for cell in sheet[1] if cell.value and str(cell.value).strip() != ""]
                elif extension_user == '.xls':
                    hojas = workbook.sheet_by_name(hoja)
                    #row_title = hojas.row_values(0)
                    row_title = [value for value in hojas.row_values(0) if value and str(value).strip() != ""]
                else:
                    raise ValueError("Formato de archivo no soportado: usa .xls o .xlsx")
                
                print(f"Columnas de la hoja {hoja}",row_title)
                # VALIDA TODAS LAS COLUMAS PARA CADA HOJA
                all_columns_validate_pivot = True
                for column in row_title:
                    sheet_column_name = column.lower()
                    validate_name_column = name_validate_sql(sheet_column_name)
                    if validate_name_column[0]:
                        print(f"     >>> Column OK -> {column}")
                    else:
                        all_columns_validate_pivot = False
                        #AGREGA EL ERROR DE LA VALIDADCION DE NOMBRES PARA LA COLUMNA DE CADA HOJA
                        column_name_error.append(f"SHEET [ {hoja} ]: {validate_name_column[1]}")
                #SOLO SI TODAS LAS COLUMNAS DE LA HOJA PASARON EL CONTROL SE AGREGA A LA LISTA DE HOJAS A PROCESAR
                if all_columns_validate_pivot:
                    SHEET_OK_INSERT.append(hoja)
            else:
                #AGREGA EL MENSAJE DE ERROR DE LA VALIDACION DE NOMBRE PARA LA HOJA
                table_error_name.append(validate_name[1])
        
        #Muestra mensaje de alerta con las tablas con las que se tuvo problemas
        message_error= ""
        if len(table_error_name):
            message_error = "NOMBRES DE HOJAS\n"
            for i in table_error_name:
                message_error = message_error+f"\n - {i}"

        if len(column_name_error):
            message_error = message_error+"\nNOMBRES DE COLUMNAS\n"
            for i in column_name_error:
                message_error = message_error+f"\n - {i}"
        if len(column_name_error) or len(table_error_name): 
            if messagebox.askyesno("ALERTA - Hojas no aptas", f"{message_error}\nLas hojas mencionadas no se tomaran en cuenta\n¿Desea continuar?"):
                SQL_INSERT_DATA(SHEET_OK_INSERT,[])
            else:
                messagebox.showinfo("RECOMENDACIÓN", "Corrige las alertas mostradas y vuelve a intentarlo")
        else:
            SQL_INSERT_DATA(SHEET_OK_INSERT,[])
    else:

        column_name_error = []
        if len(sheet_name)>0 and len(t_name)>0:
            print("Tabla lista para ser insertada",t_name)
            #SELECCIONA LA HOJA ACTUAL EN ITERACION Y OBTIENE LA CABECERA O PRIMERA FILA

            if extension_user == '.xlsx':
                sheet = workbook[sheet_name]
                # Leer la primera fila (títulos)
                #row_title = [cell.value for cell in sheet[1]]
                row_title = [cell.value for cell in sheet[1] if cell.value and str(cell.value).strip() != ""]
            elif extension_user == '.xls':
                hojas = workbook.sheet_by_name(sheet_name)
                #row_title = hojas.row_values(0)
                row_title = [value for value in hojas.row_values(0) if value and str(value).strip() != ""]
            else:
                raise ValueError("Formato de archivo no soportado: usa .xls o .xlsx")
            print(f"Columnas de la hoja {sheet_name}",row_title)
            # VALIDA TODAS LAS COLUMAS PARA CADA HOJA
            all_columns_validate_pivot = True

            #EXTRAE DATOS DE LAS ENTRADAS DE TEXTO
            for widget in f_columns.winfo_children():
                if isinstance(widget, Entry):
                    # print("HOLA",widget.get())
                    sheet_column_name = widget.get().lower()
                    validate_name_column = name_validate_sql(sheet_column_name)
                    if validate_name_column[0]:
                        print(f"     >>> Column OK -> {sheet_column_name}")
                    else:
                        all_columns_validate_pivot = False
                        #AGREGA EL ERROR DE LA VALIDADCION DE NOMBRES PARA LA COLUMNA DE CADA HOJA
                        column_name_error.append(f"SHEET [ {sheet_name} ]: {validate_name_column[1]}")
            if all_columns_validate_pivot:
                print("Listo para insertar")
                SQL_INSERT_DATA(
                    [sheet_name], #PESTAÑAS DEL EXCEL
                    [t_name] # NOMBRE PERSONALIZADO DE LA TABLA
                )
            else:
                message_error = ""
                if len(column_name_error):
                    message_error = message_error+"\nNOMBRES DE COLUMNAS\n"
                    for i in column_name_error:
                        message_error = message_error+f"\n - {i}"
                        
                    messagebox.showinfo("ALERTA", f"{message_error}\n Los nombres de columnas anteriores deben ser corregidos" )
        else:
                if len(sheet_name) == 0:
                    messagebox.showwarning("Alerta", "Debe seleccionar una Hoja")
                elif len(t_name) == 0:
                    messagebox.showwarning("Alerta", "Debe asignar un nombre como tabla")
            # print(mensaje)  # Mostrar el mensaje de error si no es válido
def open_file_excel(file_path):
    global workbook
    global row_title
    global row_muestra
    global f_columns
    global extension_user

    extension = os.path.splitext(file_path)[1].lower()

    hojas = []
    row_title = []
    row_muestra = []

    if extension == '.xlsx':
        extension_user = ".xlsx"
        workbook = load_workbook(file_path)
        hojas = workbook.sheetnames
        print("Hojas en el libro:")
        for hoja in hojas:
            print(hoja.lower())
        list_sheets.config(values=hojas)
        if hojas:
            list_sheets.set(hojas[0])
            ws = workbook[hojas[0]]
            #row_title = [cell.value for cell in ws[1]]
            row_title = [cell.value for cell in ws[1] if cell.value and str(cell.value).strip() != ""]
            row_muestra = [cell.value for cell in ws[2]]

    elif extension == '.xls':   
        extension_user = ".xls"
        workbook = xlrd.open_workbook(file_path)
        hojas = workbook.sheet_names()
        tabs = [i for i in hojas]
        # tabs = [i.lower() for i in hojas]
        print("Hojas en el libro:")
        for hoja in tabs:
            print(hoja.lower())
        list_sheets.config(values=tabs)
        if tabs:
            list_sheets.set(tabs[0])
            ws = workbook.sheet_by_name(tabs[0])
            row_title = ws.row_values(0)
            row_muestra = ws.row_values(1)
            print("MUESTRA: ",row_muestra)

    else:
        raise ValueError("Formato de archivo no soportado: usa .xls o .xlsx")


    for widget in f_columns.winfo_children():
        widget.destroy()
    # Mostrar los títulos como etiquetas y entradas
    rows = 0
    columns = 1
    for i in row_title:
        Label(f_columns, text=f"{i} :").grid(row=rows, column=columns - 1, sticky='e')
        Entry(f_columns).grid(row=rows, column=columns)
        rows += 1
        if rows == 6:
            rows = 0
            columns += 2

    # Asignar valores por defecto a los Entry
    rows = 0
    for widget in f_columns.winfo_children():
        if isinstance(widget, Entry):
            widget.insert(0, row_title[rows] if rows < len(row_title) else "")
            rows += 1
    BTN_CHARGE_DATA.config(state="normal")
    #EXTRAE DATOS DE LAS ENTRADAS DE TEXTO
    # for widget in f_columns.winfo_children():
    #     if isinstance(widget, Entry):
    #         print("Column: ", widget.get())
def get_file_path():
    global name_path_file  # Declaración global para usar la variable fuera de la función
    file = filedialog.askopenfilename(
        title="Selecciona un archivo Excel",
        filetypes=[
            ("Archivos compatibles", "*.xls *.xlsx *.csv"),  # <- filtro combinado
            ("Archivos Excel (.xls)", "*.xls"),
            ("Archivos Excel (.xlsx)", "*.xlsx"),
            ("Archivos CSV", "*.csv"),
            ("Todos los archivos", "*.*")
        ]
    )

    if file:
        print(f"Archivo seleccionado: {file}")
        lbl_file_path.delete(0, tk.END)      # Borrar todo el contenido
        lbl_file_path.insert(0, file)  # Insertar texto desde el inicio
        name_path_file = file  # Asignación de la variable global
        open_file_excel(file)

def get_database_info(tree):
    # Limpiar todos los elementos
    for item in tree.get_children():
        tree.delete(item)
    # Conectar a la base de datos
    base_dir = os.path.dirname(os.path.abspath(__file__))
    # db_path = os.path.join(base_dir, "DATA", "data_main.db")
    # conn = sqlite3.connect(db_path)
    conn = sqlite3.connect("DATA/data_main.db")
    cursor = conn.cursor()

    # Ejecutar la consulta para obtener las tablas
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")

    # Obtener los nombres de las tablas
    tablas = cursor.fetchall()

    data_main = []
    for tabla in tablas:
        data=[]
        # Consultar la estructura de la tabla
        cursor.execute(f"PRAGMA table_info({tabla[0]});")
        # Obtener los nombres de las columnas
        columnas = cursor.fetchall()
        # sum_columns = len(columnas)
        for j in columnas:
            #(9, 'name_column', 'TEXT')
            # data.append((j[0],j[1],j[2]))
            data.append((str(j[0]),j[1]))
        data_main.append((tabla[0],f"{len(columnas)} columnas",data))
    # return data_main
    # print
    # Insertar nodos padres e hijos dinámicamente
    for nombre, rows, hijos, *tag in data_main:
        tag = tag[0] if tag else "padre"  # Si hay un tag especial, lo usa; de lo contrario, "padre"
        padre_id = tree.insert("", "end", values=(nombre, rows), tags=(tag,))
        
        for item, fecha in hijos:
            tree.insert(padre_id, "end", values=(item, fecha, ""), tags=("hijo",))


def execute_query(data_text):
    global DATA_EXECUTE
    def ejecutar_bloque(conexion, bloque_sql, limite_select=10000):
        """
        Ejecuta un bloque con múltiples sentencias SQL.
        Devuelve una lista de resultados (uno por sentencia):
        - SELECT/PRAGMA → [sql, filas, columnas]
        - DML (INSERT, UPDATE, DELETE, REPLACE) → [sql, rowcount]
        - DDL (CREATE, DROP, ALTER) → [sql, True]
        - Error → [sql, "Error", mensaje_error]
        """
        cursor = conexion.cursor()
        resultados = []

        # Separar sentencias por ;
        sentencias = [s.strip() for s in bloque_sql.split(";") if s.strip()]
        print("THIS IS A SENTENCES:",sentencias)
        for sql in sentencias:
            qtype = sql.split()[0].upper()
            print("THIS IS TYPE",qtype)
            try:
                cursor.execute(sql)

                if qtype in ("SELECT", "PRAGMA","WITH"):
                    # filas = cursor.fetchall()
                    filas = cursor.fetchmany(limite_select)
                    columnas = [desc[0] for desc in cursor.description] if cursor.description else []
                    resultados.append([sql, filas, columnas])
                elif qtype in ("INSERT", "UPDATE", "DELETE", "REPLACE"):
                    resultados.append([sql, cursor.rowcount])
                else:
                    resultados.append([sql, True])  # CREATE, DROP, ALTER, etc.
            except Exception as e:
                resultados.append([sql, "Error", str(e)])

        return resultados

    for item in TREE_item.get_children():
        TREE_item.delete(item)
    # print("DELETE COMMENT",data_text.split("\n"))
    query_no_comments=""
    for i in data_text.split("\n"):
        if i[0:2] != "--":
            query_no_comments +=  F" {i}"
        else:
            print(f"THIS IS A COMMENT: {i}")
    print("COMMENTS CLEANED",query_no_comments)
    query = query_no_comments.replace("\n", " ").replace("\t", " ")
    conn = sqlite3.connect("DATA/data_main.db")
    DATA_EXECUTE = ejecutar_bloque(conn,query)
    line_query = 1
    for i in DATA_EXECUTE:
        # Definir estilos de fila alternos usando tags
        TREE_item.tag_configure('evenrow', background='#f0f0f0')  # gris claro
        TREE_item.tag_configure('oddrow', background='white')     # blanco
        TREE_item.tag_configure('error', background='#FF7659')     # blanco
        tag = 'evenrow' if line_query % 2 == 0 else 'oddrow'
        if i[1] == "Error":
            tag = 'error'
        TREE_item.insert("", "end", values=(line_query, f"{f'[{i[1]}] {i[2]} in {i[0]}' if i[1] == 'Error' else F'[OK] {i[0]}'}"),tags=(tag,)) #, tags=(tag,)
        line_query+=1
        conn.close()
root = Tk()
root.title("SQLXEL Administrator v1.0.0 Powered by Breyner J")
root.iconbitmap("SOURCE/icon.ico")
Label(root, text="SQLXEL Administrator", font=("Arial", 12, "bold")).pack(fill=X)

F_main = Frame(root, bg="green")
F_main.pack(fill=BOTH)

F_filter = LabelFrame(F_main, text="Filter and selection")
F_filter.pack(fill=BOTH, expand=True,side=LEFT)
Label(F_filter,text="Origin").pack(expand=True, fill=X)
lbl_file_path = Entry(F_filter)
lbl_file_path.pack(expand=True, fill=X)
Label(F_filter,text="Destination").pack(expand=True, fill=X)
lbl_file_path_out = Entry(F_filter)
lbl_file_path_out.pack(expand=True, fill=X)
#Button(F_filter, text="Open file", bg="green",fg="white",font=("arial",12,"bold"), command=lambda: get_file_path()).pack()

FR_input = Frame(F_filter)
FR_input.pack()

Label(FR_input, text="Select sheet :").grid(        row=0, column=0, sticky='e')
Label(FR_input, text="Name as table :").grid(       row=0, column=2, sticky='e')
Label(FR_input, text="All of the sheets :").grid(   row=0, column=4, sticky='e')
Label(FR_input, text="Suffix :").grid(              row=0, column=6, sticky='e')
Label(FR_input, text="Save locally :").grid(        row=0, column=8, sticky='e')

# === Canvas + Scrollbar en el contenedor principal ===
wrapper = Frame(F_filter)
wrapper.pack(fill="both", expand=True)

canvas = tk.Canvas(wrapper, height=120)
canvas.pack(side="top", fill="both", expand=True)

scroll_x = tk.Scrollbar(wrapper, orient="horizontal", command=canvas.xview)
scroll_x.pack(side="bottom", fill="x")
canvas.configure(xscrollcommand=scroll_x.set)

# === f_columns está completamente vacío y será usado como contenedor temporal ===
f_columns = tk.Frame(canvas)  # <<<<<< ESTE es el frame vacío
canvas.create_window((0, 0), window=f_columns, anchor="n")

# === Ajustar el área de scroll cuando f_columns cambie ===
def actualizar_scroll(event):
    canvas.configure(scrollregion=canvas.bbox("all"))
f_columns.bind("<Configure>", actualizar_scroll)
def al_seleccionar_hoja(event):
        BTN_CHARGE_DATA.config(state="normal")
        for widget in f_columns.winfo_children():
            widget.destroy()
        # f_columns = Frame(FR_content_frame)
        # f_columns.grid(row=4, column=0, columnspan=2)
        seleccion = list_sheets.get()  # Obtener el valor seleccionado

        if extension_user == '.xlsx':
            sheet = workbook[seleccion]
            # Leer la primera fila (títulos)
            #first_line = [cell.value for cell in sheet[1]]
            first_line = [cell.value for cell in sheet[1] if cell.value and str(cell.value).strip() != ""]
        elif extension_user == '.xls':
            hojas = workbook.sheet_by_name(seleccion)
            #first_line = hojas.row_values(0)
            first_line = [value for value in hojas.row_values(0) if value and str(value).strip() != ""]
        else:
            raise ValueError("Formato de archivo no soportado: usa .xls o .xlsx")
        
        rows = 0
        columns = 1
        for i in first_line:
            Label(f_columns, text=f"{i} :").grid(row=rows, column=columns - 1, sticky='e')
            Entry(f_columns).grid(row=rows, column=columns)
            rows += 1
            if rows == 6:
                rows = 0
                columns += 2
        # Asignar valores por defecto a los Entry
        rows = 0
        for widget in f_columns.winfo_children():
            if isinstance(widget, Entry):
                widget.insert(0, first_line[rows] if rows < len(first_line) else "")
                rows += 1
# Crear el Combobox
list_sheets = ttk.Combobox(FR_input,state="readonly")
list_sheets.grid(row=0, column=1, sticky='w')
list_sheets.bind("<<ComboboxSelected>>", al_seleccionar_hoja)
table_name = Entry(FR_input)
table_name.grid(row=0, column=3, sticky='we')
sufijo_name = Entry(FR_input)
sufijo_name.grid(row=0, column=7, sticky='we')

def cambiar_texto():
    # Cambiar el texto del Checkbutton según su estado
    if var.get():
        checkbutton.config(text="Yes")  # Si está seleccionado, mostrar "Sí"
        list_sheets.config(state='disabled')
        table_name.config(state='disabled')
        BTN_CHARGE_DATA.config(state="normal")
        try:
            for widget in f_columns.winfo_children():
                widget.configure(state='disabled')
        except:
            pass  # Si el widget no tiene la opción 'state'
    else:
        checkbutton.config(text="No")   # Si no está seleccionado, mostrar "No"
        list_sheets.config(state='readonly')
        table_name.config(state='normal')
        try:
            for widget in f_columns.winfo_children():
                widget.configure(state='normal')
        except:
            pass  # Si el widget no tiene la opción 'state'
def cambiar_texto_save():
    # Cambiar el texto del Checkbutton según su estado
    if var_saved.get():
        checkbutton_save.config(text="Yes")  # Si está seleccionado, mostrar "Sí"
    else:
        checkbutton_save.config(text="No")   # Si no está seleccionado, mostrar "No"
        
        

def refrescar():
    checkbutton.config(text="No")   # Si no está seleccionado, mostrar "No"
    var.set(False)
    list_sheets.config(state='readonly')
    table_name.config(state='normal')
    # Muestra la notificación
    try:
        
        for widget in f_columns.winfo_children():
            widget.destroy()
        f_columns.update_idletasks()  # Fuerza la actualización de la interfaz
    except:
            pass  # Si el widget no tiene la opción 'state'
    table_name.delete(0, END)
    sufijo_name.delete(0, END)
    BTN_CHARGE_DATA.config(state="normal")
    list_sheets.set('')
    get_database_info(tree)

# Variable para almacenar el estado del Checkbutton
var = BooleanVar()
var_saved = BooleanVar()
# Crear el Checkbutton
checkbutton = Checkbutton(FR_input, text="No", variable=var, command=cambiar_texto)
checkbutton.grid(row=0, column=5, sticky='w')

checkbutton_save = Checkbutton(FR_input, text="No", variable=var_saved, command=cambiar_texto_save)
checkbutton_save.grid(row=0, column=9, sticky='w')

FR_BTN_main = Frame(FR_input)
FR_BTN_main.grid(row=1, column=0, columnspan=9) 

Button(FR_BTN_main, text="Open file", bg="green",fg="white",font=("arial",12,"bold"), command=lambda: get_file_path()).grid(row=0, column=0)
BTN_CHARGE_DATA=Button(FR_BTN_main, text="Data load",bg="lightgreen",fg="black",font=("arial",12,"bold"), command=lambda: save_data(name_path_file, table_name.get(), list_sheets.get(),var.get(),sufijo_name.get()))
BTN_CHARGE_DATA.grid(row=0, column=1)

BTN_REFRESH=Button(FR_BTN_main, text="Refresh", bg="lightblue",fg="black",font=("arial",12,"bold"), command=lambda: refrescar())
BTN_REFRESH.grid(row=0, column=2)

F_database = LabelFrame(F_main,text="Database")
F_database.pack(side=RIGHT,fill=BOTH)

# Estilos para Treeview
style = ttk.Style()
style.configure("Treeview", rowheight=25)  # Ajustar altura de filas




# Crear un Treeview con columnas
tree = ttk.Treeview(F_database, columns=("Table", "Columns"), show="headings", height=5)

# Definir encabezados
tree.heading("Table", text="Table")
tree.heading("Columns", text="Columns")

# Configurar ancho de columnas
tree.column("Table", width=190, anchor="w")
tree.column("Columns", width=100, anchor="center")

# Definir colores con tags
tree.tag_configure("padre", background="lightblue")   # Rojo claro
tree.tag_configure("hijo", background="#ccffcc")    # Verde claro
tree.tag_configure("especial", background="#ccccff")  # Azul claro

get_database_info(tree)
# Posicionar el Treeview
tree.pack(expand=True, fill="both")

frm_query = Frame(root, bg="red")
frm_query.pack(fill=BOTH)

# Función para marcar palabras específicas
def marcar_palabras(event=None):
    # Limpiar cualquier formato anterior
    texto.tag_remove("resaltado", "1.0", END)
    
    # Obtener el contenido del texto
    texto_content = texto.get("1.0", END)
    
    # Recorrer cada palabra a resaltar
    for palabra in palabras_reservadas:
        # Usar expresión regular para encontrar solo palabras completas (ignorando mayúsculas/minúsculas)
        pattern = r'\b' + re.escape(palabra) + r'\b'  # \b asegura que sea una palabra completa
        matches = list(re.finditer(pattern, texto_content, re.IGNORECASE))  # Añadir re.IGNORECASE
        
        # Marcar cada coincidencia
        for match in matches:
            start_pos = texto.index(f"1.0 + {match.start()} chars")  # Convertir la posición a formato Tkinter
            end_pos = texto.index(f"1.0 + {match.end()} chars")  # Calcular el final
            texto.tag_add("resaltado", start_pos, end_pos)  # Agregar el tag para resaltar

    # Aplicar el color al tag "resaltado"
    texto.tag_configure("resaltado", foreground="orange")  # Cambiar el color a rojo
    
    ultima_linea = [int(item_numbers.index('end-1c').split('.')[0]), int(texto.index('end-1c').split('.')[0])]  # 'end-1c' elimina el carácter de nueva línea al final
    # Dividir la última línea para obtener el número de línea
    print(ultima_linea)
    if ultima_linea[1] != ultima_linea[0]:
        item_numbers.config(state="normal")
        item_numbers.delete(2.0, END)
        for i in range(2,ultima_linea[1]+1):
            item_numbers.insert(END,f"\n{i}")
        item_numbers.config(state="disabled")
    # elif ultima_linea[1]<ultima_linea[0]:
    #     pass
    # numero_de_lineas = int(ultima_linea.split('.')[0])



    # for i in range(1,45):
    #     item_numbers.insert(END, f"{i}\n")
# def line_enter(event=None):
#     print("presiono enter")
# def line_delete(event=None):
#     print("presiono delete")
# Crear el widget Text
F_editor = Frame(frm_query)
F_editor.pack(fill=BOTH)

item_numbers = Text(F_editor, height=15,width=4, wrap="none",bg='black',font=("Courier",11,"bold"),fg='green',insertbackground='white')  # wrap="word" evita cortar palabras
item_numbers.pack(side=LEFT, fill=X)

texto = Text(F_editor, height=15, bg='black',wrap="none",font=("Courier",11,"bold"),fg='white',insertbackground='white')  # wrap="word" evita cortar palabras
texto.pack(side=LEFT, fill=X, expand=True)
texto.focus_set()

# Asociar la función al evento de escribir (KeyRelease)
texto.bind("<KeyRelease>", marcar_palabras)
# texto.bind("<Return>", line_enter)
# texto.bind("<Delete>", line_delete)
scrollbar = Scrollbar(F_editor)
scrollbar.pack(side=RIGHT, fill=Y)

# Configurar los widgets Text para que usen el mismo Scrollbar
item_numbers.config(yscrollcommand=scrollbar.set)
texto.config(yscrollcommand=scrollbar.set)

# Configurar el Scrollbar para que maneje el desplazamiento de ambos widgets
scrollbar.config(command=lambda *args: [item_numbers.yview(*args), texto.yview(*args)])
# Función para permitir que ambos Text se desplacen con la rueda del ratón
def on_mouse_wheel(event):
    delta = -1 * (event.delta // 120)
    # Si el evento se genera en el widget texto, desplazamos ambos
    if event.widget == texto:
        item_numbers.yview_scroll(int(delta), "units")
    elif event.widget == item_numbers:
        texto.yview_scroll(int(delta), "units")

# Asociar el evento de rueda del ratón en ambos widgets
texto.bind_all("<MouseWheel>", on_mouse_wheel)
item_numbers.bind_all("<MouseWheel>", on_mouse_wheel)
item_numbers.insert(END, "1")
item_numbers.config(state="disabled")

Button(root,text="Execute query",font=("Arial",12,"bold"), cursor="hand2",background="orange",fg="black",command=lambda:execute_query(texto.get("1.0", END))).pack()
frm_response_iten = Frame(root,bg="grey")
frm_response_iten.pack(fill=BOTH)
# Crear un Treeview con columnas
TREE_item = ttk.Treeview(frm_response_iten, columns=("item", "Query"), show="headings", height=3)

# Definir encabezados
TREE_item.heading("item", text="Item")
TREE_item.heading("Query", text="Query")
# Vincular evento de selección
def on_row_selected(event):
    def show_data(columns, data):
        for widget in frm_response_show.winfo_children():
            widget.destroy()

        # Crear el Treeview
        tree = ttk.Treeview(frm_response_show, columns=tuple(columns), show="headings")
        tree.pack(expand=True, fill='both', padx=10, pady=10)

        # Configurar encabezados
        for col in columns:
            tree.heading(col, text=col)

        # Definir estilos de fila alternos usando tags
        tree.tag_configure('evenrow', background='#f0f0f0')  # gris claro
        tree.tag_configure('oddrow', background='white')     # blanco

        # Insertar datos con colores alternos
        for idx, row in enumerate(data):
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
            tree.insert("", "end", values=row, tags=(tag,))
    # Obtener referencia al widget
    tree = event.widget  
    
    # Obtener el item seleccionado
    selected_item = tree.selection()
    if selected_item:
        item = selected_item[0]
        valores = tree.item(item, "values")
        print(f"Fila seleccionada: {valores[0]}",DATA_EXECUTE[int(valores[0])-1])
        
        if len(DATA_EXECUTE[int(valores[0])-1]) == 3 and DATA_EXECUTE[int(valores[0])-1][1] != "Error":
            
            for i in DATA_EXECUTE[int(valores[0])-1]:
                print(i)
            show_data(DATA_EXECUTE[int(valores[0])-1][2],DATA_EXECUTE[int(valores[0])-1][1])
TREE_item.bind("<<TreeviewSelect>>", on_row_selected)
def on_resize(event):
    ancho = root.winfo_width()
    TREE_item.column("item", width=int(ancho * 0.05), anchor="w")
    TREE_item.column("Query", width=int(ancho * 0.95), anchor="w")
root.bind("<Configure>", on_resize)

# Definir colores con tags
# TREE_item.tag_configure("padre", background="#ffcccc")   # Rojo claro
# TREE_item.tag_configure("hijo", background="#ccffcc")    # Verde claro
# TREE_item.tag_configure("especial", background="#ccccff")  # Azul claro

# get_database_info(tree)
# Posicionar el Treeview
TREE_item.pack(expand=True, fill="both")

frm_response_show = Frame(root, bg="#f0f0f0")
frm_response_show.pack(fill=BOTH, expand=True)
print(r"""
               _____   ____   __     __   __ ______ __      __ 
              / ____  / __ \ | |     \ \ / /|  ____| |     | | This is a aplication 
             | (____ | |  | || |      \ V / | |__  | |     | | desingned for convert 
             \____  \| |  | || |       > <  |  __| | |     | | Excel files in tables 
              ____) )| |__| || |____  / . \ | |____| |____ |_| of the database whit 
             |_____/  \____/ |______|/_/ \_\|______|______|(_) powered by SQLite.
                        \_\_
                            # Your App is ready #
                                        Developed by Breyner J.
""")
root.mainloop()