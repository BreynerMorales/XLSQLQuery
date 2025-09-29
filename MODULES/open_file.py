import os
from openpyxl import load_workbook
import xlrd
from tkinter import Label, Entry

from MODULES.vars import varsco
def open_file_excel(file_path,list_sheets,f_columns):
    # global varsco["workbook"]
    # global varsco["row_title"]
    # global row_muestra
    # global f_columns
    # global extension_user

    extension = os.path.splitext(file_path)[1].lower()

    hojas = []
    row_title = []
    # row_muestra = []

    if extension == '.xlsx':
        varsco["extension_user"] = ".xlsx"
        varsco["workbook"] = load_workbook(file_path)
        hojas = varsco["workbook"].sheetnames
        print("Hojas en el libro:")
        for hoja in hojas:
            print(hoja.lower())
        list_sheets.config(values=hojas)
        if hojas:
            list_sheets.set(hojas[0])
            ws = varsco["workbook"][hojas[0]]
            #row_title = [cell.value for cell in ws[1]]
            varsco["row_title"] = [cell.value for cell in ws[1] if cell.value and str(cell.value).strip() != ""]
            varsco["row_muestra"] = [cell.value for cell in ws[2]]

    elif extension == '.xls':   
        varsco["extension_user"] = ".xls"
        varsco["workbook"] = xlrd.open_workbook(file_path)
        hojas = varsco["workbook"].sheet_names()
        tabs = [i for i in hojas]
        # tabs = [i.lower() for i in hojas]
        print("Hojas en el libro:")
        for hoja in tabs:
            print(hoja.lower())
        list_sheets.config(values=tabs)
        if tabs:
            list_sheets.set(tabs[0])
            ws = varsco["workbook"].sheet_by_name(tabs[0])
            varsco["row_title"] = ws.row_values(0)
            varsco["row_muestra"] = ws.row_values(1)
            print("MUESTRA: ",varsco["row_muestra"])

    else:
        raise ValueError("Formato de archivo no soportado: usa .xls o .xlsx")


    for widget in f_columns.winfo_children():
        widget.destroy()
    # Mostrar los títulos como etiquetas y entradas
    rows = 0
    columns = 1
    for i in varsco["row_title"]:
        Label(f_columns, text=f"{i} :").grid(row=rows, column=columns - 1, sticky='e')
        Entry(f_columns).grid(row=rows, column=columns)
        rows += 1
        if rows == 6:
            rows = 0
            columns += 2

    # Asignar valores por defecto a los Entry
    rows = 0
    for widget in f_columns.winfo_children():
        if isinstance(widget, Entry):
            widget.insert(0, varsco["row_title"][rows] if rows < len(varsco["row_title"]) else "")
            rows += 1
    